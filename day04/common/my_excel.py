"""
    对Excel是读写操作封装
"""
from openpyxl import load_workbook

class MyExcel:

    def __init__(self,excel_path,sheet_name):
        # 加载Excel工作簿，得到工作簿 workbook
        wb = load_workbook(excel_path)
        # 选择一个表单
        self.sh = wb[sheet_name]

    def read_data(self):
        # 读取Excel中的数据
        data = list(self.sh.values)
        print(data)
        header = data[0]#获取列名
        #存储表单下读取到的所有数据 - 每个成员都是一个字典
        all_data = []
        for row in data[1:]:
            row_dict = dict(zip(header, row))
            all_data.append(row_dict)
        return all_data

if __name__=='__main__':
    #文件路径
    excel_path = r"D:\coderesource\python_basic04\day02\testdatas\测试用例v1.xlsx"
    wb = load_workbook(excel_path)
    me = MyExcel(excel_path,"注册接口")

    cases = me.read_data()
    for case in cases:
        print(case)



