import openpyxl


from openpyxl import load_workbook

#加载Excel工作簿，得到工作簿 workbook
wb = load_workbook(r"D:\coderesource\python_basic04\day02\testdatas\测试用例v1.xlsx")

#选择一个表单
sh = wb["注册接口"]

#根据行号和列号进行读取数据
cell_vale = sh.cell(3,4).value
print(cell_vale)

#读取所有的行
# for row in sh.rows:
#     for item in row:
#         print(item.value,end=" ")
#     print()

#得到sheet的总行号，总列号
row_nums = sh.max_row
col_nums = sh.max_column

#只读取第一行作为key
#行号是1
keys = []
for col_index in range(1,col_nums + 1):
    keys.append(sh.cell(1,col_index).value)
print(keys)

for row_index in range(2,sh.max_row + 1):
    values = []
    for col_index in range(1,sh.max_column + 1):
        values.append(sh.cell(row_index,col_index).value)
    #keys 与 values 打包 - zip函数
    case = dict(zip(keys,values))
    print(case)


#使用第二种方法进行处理Excel
data = list(sh.values)
print(data)
header = data[0]
all_data = []
for row in data[1:]:
    row_dict = dict(zip(header, row))
    print(row_dict)




#列表推导式 列表名 = [值 表达式]
#keys = [sh.cell(1,col_index).value for col_index in range(1,col_nums + 1)]


